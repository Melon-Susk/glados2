from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pytz


class Util:
    def __init__(self):
        pass
    
    @staticmethod
    def reset(driver):
        for i in range(5):
            try:
                profileMenu = driver.find_element(By.XPATH, '//*[text()="Profil"]')
                profileMenu.click()
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[text()="Spielerübersicht"]')))
                profileMenu.click()
                break
            except:
                print(f"Reset fail {i+1}")
                try:
                    driver.find_element(By.CSS_SELECTOR, '.icon.icon-tutorial.icon-close-button').click()
                except:
                    pass
                time.sleep(float(i))
                continue
        try:
            WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[text()="Spielerübersicht"]')))
            return False
        except:
            return True
        
    @staticmethod
    def checkNewDay(lastCheck):
        jetzt = datetime.now()

        if jetzt.date() > lastCheck.date():
            ist_neuer_tag = True
        else:
            ist_neuer_tag = False

        return ist_neuer_tag
    
    @staticmethod
    def loadJsonToDict(dateipfad):
        try:
            with open(dateipfad, 'r', encoding='utf-8') as datei:
                daten = json.load(datei)
                return daten
        except FileNotFoundError:
            print("Die Datei wurde nicht gefunden:", dateipfad)
            return None
        except json.JSONDecodeError:
            print("Fehler beim Parsen der JSON-Datei:", dateipfad)
            return None
    
    @staticmethod
    def loadDatetimeJsonTodict(dateiname):
        with open(dateiname, 'r') as file:
            daten = json.load(file)
        
        # Konvertiere die String-Datumsangaben zurück in datetime-Objekte
        for key, value in daten.items():
            try:
                daten[key] = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                # Falls das Datum nicht im erwarteten Format ist, behalte den String bei
                pass
        
        return daten
    
    @staticmethod
    def appendToOverviewJson(dateipfad, accountName, castleName, resourceDict):
        with open(dateipfad, 'r', encoding='utf-8') as file:
            file_content = file.read()
            if not file_content:
                daten = {}
            else:
                daten = json.loads(file_content)
        
        if accountName not in daten:
            daten[accountName] = {}

        daten[accountName][castleName] = resourceDict

        with open(dateipfad, 'w', encoding='utf-8') as file:
            json.dump(daten, file)
    
    @staticmethod
    def determineCastleAndSilverAmount(excel_path='silverOverview.xlsx'):
        # JSON-Daten laden
        with open('resourceOverview.json', 'r') as file:
            data = json.load(file)

        # Liste für die Ergebnisse
        ergebnisse = []

        # Verarbeite jeden Top-Level-Schlüssel
        for email, burgen in data.items():
            anzahl_burgen = len(burgen)
            summe_silber = sum(burg.get("Silber", 0) for burg in burgen.values())
        
            # Benötigtes Silber berechnen
            ben_silber = anzahl_burgen * 1000
            if ben_silber <= summe_silber:
                ben_silber = "SILBER AUSREICHEND"
            else:
                ben_silber = ben_silber - summe_silber

            # Ergebnisse hinzufügen
            ergebnisse.append({"E-Mail": email, "Anzahl Burgen": anzahl_burgen, "Summe Silber": summe_silber, "Noch benötigt": ben_silber})

        # Erstelle einen DataFrame aus den Ergebnissen
        df = pd.DataFrame(ergebnisse)
        df.to_excel(excel_path, index=False, engine='openpyxl')
    
    @staticmethod
    def autoAdjustColumnWidths(excel_path='silverOverview.xlsx'):
        # Lade das Workbook
        wb = load_workbook(excel_path)
        
        # Gehe durch alle Worksheets im Workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Durchlaufe alle Spalten im Worksheet
            for col in ws.columns:
                max_length = 0
                column = col[0].column   # Hole die Spaltenbuchstaben/nummer
                
                # Bestimme die maximale Länge des Inhalts in der Spalte
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                
                # Passe die Spaltenbreite an (füge 2 hinzu für ein bisschen Puffer)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    
        # Speichere die Änderungen im Workbook
        wb.save(excel_path)
    
    @staticmethod
    def isNight():
        timezone_de = pytz.timezone('Europe/Berlin')
        jetzt = datetime.now(timezone_de)

        if jetzt.hour >= 23:
            return True
        elif jetzt.hour < 7:
            return True
        return False
    
    @staticmethod
    def isEvening():
        timezone_de = pytz.timezone('Europe/Berlin')
        jetzt = datetime.now(timezone_de)

        if jetzt.hour > 19:
            return True
        return False
    
    @staticmethod
    def refreshDatetimeDict(datetime_dict, castlename_dict):
        # Überprüfe jeden Schlüssel im referenz_dict
        for key in castlename_dict:
            # Wenn der Schlüssel im datetime_dict nicht existiert, füge ihn mit dem aktuellen Datum/Zeit hinzu
            if key not in datetime_dict:
                datetime_dict[key] = datetime.now()
        
        # Konvertiere datetime-Objekte zu Strings für JSON-Kompatibilität
        for key, value in datetime_dict.items():
            if isinstance(value, datetime):
                datetime_dict[key] = value.strftime("%Y-%m-%d %H:%M:%S")

        # Speichere das aktualisierte Dictionary als JSON-Datei
        with open('timeSortedAccounts.json', 'w') as file:
            json.dump(datetime_dict, file)

